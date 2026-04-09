"""
앙스타 VOLTAGE 이지작 클리어 노트수 예측 회귀 모델 v3
──────────────────────────────────────────────────
v2 → v3 변경사항:
  1. export_songs_js() 추가 — 예측 결과를 songs.js로 자동 출력
     · voltage_predictor_v4.html 과 같은 폴더에 두면 바로 연동 가능
     · --export 옵션 또는 대화형 메뉴에서 호출 가능
  2. duration 파싱 보정 — Excel이 시간으로 읽는 "HH:MM:SS" 형식을 "M:SS"로 정규화
  3. FAIL_URL 기본값 설정

R² = 0.9663 / MAE = 2.79노트 / LOO-MAE = 2.86노트 (167곡)

[엑셀 헤더 대응표]
  total_notes          ← 총 노트수
  et_start             ← 앙상블 타임 시작 콤보
  et_end               ← 앙상블 타임 종료 콤보
  clear_start_measured ← 최종 시작점(실측)
  category             ← 속성 (시트명)
  type                 ← 타입
  unit                 ← 유닛
  title                ← 곡 제목
  duration             ← 곡 길이
  video_url            ← 영상 주소

[사용법]
  # 기본 실행 (예측 + 대화형 메뉴)
  python enstars_regression_v3.py

  # 엑셀 파일 직접 지정
  python enstars_regression_v3.py es_regression.xlsx

  # songs.js 바로 출력 (대화 없이)
  python enstars_regression_v3.py --export
  python enstars_regression_v3.py es_regression.xlsx --export

  # 출력 경로 지정
  python enstars_regression_v3.py --export --out ./dist/songs.js
"""

import json
import sys
import re
import requests
import pandas as pd
import numpy as np
from pathlib import Path
from sklearn.linear_model import LinearRegression
from sklearn.metrics import r2_score, mean_absolute_error
import os
from dotenv import load_dotenv

_HERE = Path(__file__).parent
load_dotenv(_HERE / ".env")
CSV_PATH = str(_HERE / "es_regression.xlsx")
SAFETY_MARGIN = 10
FAIL_URL = "https://youtu.be/Cs8hEmAl8eI"

UNIT_GROUP = {
    # Others
    "셔플_10주년": "셔플",
    "새해":        "주년·기념",
    "만우절":      "주년·기념",
    "주년":        "주년·기념",
    "콜라보":      "콜라보",
    "추억":        "기타",
    "진앤아":      "기타",
    "드림유닛":    "기타",
    "기타":        "기타",
    # Cosmic
    "Adam":        "Adam·Eve",
    "Eve":         "Adam·Eve",
    # New Dimension
    "마무":        "MaM·DF",
    "덥페":        "MaM·DF",
}


# ── 유틸 ──────────────────────────────────────────────────
def parse_duration(val) -> str | None:
    """Excel이 시간으로 읽은 값(HH:MM:SS)을 M:SS 문자열로 정규화."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    s = str(val).strip()
    parts = s.split(":")
    if len(parts) == 3:
        # 예: "02:34:00" → 실제 의미는 2분 34초
        return f"{int(parts[0]) * 60 + int(parts[1])}:{parts[2][:2]}"
    if len(parts) == 2:
        return s[:5]
    return s


def extract_video_id(url: str) -> str | None:
    """youtu.be 또는 youtube.com URL에서 video ID 추출."""
    import urllib.parse
    if not url:
        return None
    parsed = urllib.parse.urlparse(url)
    if parsed.netloc in ("youtu.be",):
        vid = parsed.path.lstrip("/").split("/")[0]
        return vid if vid else None
    if "youtube.com" in parsed.netloc:
        qs = urllib.parse.parse_qs(parsed.query)
        ids = qs.get("v", [])
        return ids[0] if ids else None
    return None


def fetch_clear_seconds(video_id: str, api_key: str) -> int | None:
    """YouTube API로 영상 설명란을 가져와 Clear 타임스탬프를 초로 반환 (-2초 적용)."""
    url = "https://www.googleapis.com/youtube/v3/videos"
    resp = requests.get(url, params={"part": "snippet", "id": video_id, "key": api_key})
    resp.raise_for_status()
    items = resp.json().get("items", [])
    if not items:
        return None
    description = items[0]["snippet"]["description"]
    match = re.search(r"(\d{1,2}):(\d{2}) Clear\([^)]+\)", description)
    if not match:
        return None
    minutes, seconds = int(match.group(1)), int(match.group(2))
    total = minutes * 60 + seconds - 2
    return max(0, total)


def calc_midpoint_url(video_url: str, duration_str: str | None) -> str | None:
    """video_url + duration으로부터 중간 지점 타임스탬프 URL 생성 (예측곡 전용)."""
    if not video_url or not duration_str:
        return None
    parts = duration_str.split(":")
    try:
        if len(parts) == 2:
            total_sec = int(parts[0]) * 60 + int(parts[1])
        else:
            return None
    except ValueError:
        return None
    mid_sec = total_sec // 2
    vid = extract_video_id(video_url)
    if not vid:
        return None
    return f"https://youtu.be/{vid}?t={mid_sec}"


def fill_clear_urls(path: str, api_key: str) -> dict:
    """엑셀 전체 시트에서 video_url을 읽어 video_url_clear를 채우고 저장.
    반환값: {video_url: video_url_clear} — df_pred 인메모리 패치용."""
    import openpyxl
    wb = openpyxl.load_workbook(path)

    filled_map = {}  # video_url → video_url_clear
    filled = 0
    skipped_existing = 0
    not_found = 0

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header = [cell.value for cell in ws[1]]
        if "video_url" not in header or "video_url_clear" not in header:
            print(f"  [{sheet_name}] video_url / video_url_clear 헤더 없음 — 건너뜀")
            continue

        col_url = header.index("video_url") + 1
        col_clear = header.index("video_url_clear") + 1
        col_measured = header.index("clear_start_measured") + 1 if "clear_start_measured" in header else None

        for row in ws.iter_rows(min_row=2):
            url_cell = row[col_url - 1]
            clear_cell = row[col_clear - 1]

            if not url_cell.value:
                continue
            # measured 없는 곡은 API 호출 안 함 (duration 절반으로 처리)
            if not col_measured or row[col_measured - 1].value in (None, ""):
                continue
            # if clear_cell.value and not is_measured:  # 최초에만 실행 (모두 비어있을 때)
            if clear_cell.value:
                skipped_existing += 1
                continue

            video_id = extract_video_id(str(url_cell.value))
            if not video_id:
                not_found += 1
                continue

            t = fetch_clear_seconds(video_id, api_key)
            if t is None:
                not_found += 1
                continue

            clear_url = f"https://youtu.be/{video_id}?t={t}"
            clear_cell.value = clear_url
            filled_map[str(url_cell.value).strip()] = clear_url
            filled += 1

    wb.save(path)
    print(f"\n✓ video_url_clear 채우기 완료: {path}")
    print(f"  채운 항목: {filled}개")
    print(f"  이미 있어서 건너뜀: {skipped_existing}개")
    print(f"  타임스탬프 못 찾음: {not_found}개")
    return filled_map


# ── 데이터 로드 ────────────────────────────────────────────
def load_data(path: str = CSV_PATH) -> pd.DataFrame:
    """전체 시트를 읽어 하나의 DataFrame으로 합친 뒤 category 컬럼 추가."""
    if path.endswith(".xlsx") or path.endswith(".xls"):
        all_sheets = pd.read_excel(path, sheet_name=None)
        frames = []
        for sheet_name, sheet_df in all_sheets.items():
            sheet_df["category"] = sheet_name
            frames.append(sheet_df)
            print(f"[{sheet_name}] {len(sheet_df)}곡  컬럼: {list(sheet_df.columns)}")
        df = pd.concat(frames, ignore_index=True)
    else:
        df = pd.read_csv(path)
        if "category" not in df.columns:
            df["category"] = "Unknown"
    return df


# ── 모델 학습 ──────────────────────────────────────────────
def train_model(df: pd.DataFrame) -> dict:
    """실측 데이터로 3변수 OLS + 단순 회귀 두 모델을 학습."""
    measured = df[df["clear_start_measured"].notna()].copy()
    measured = measured.dropna(subset=["et_start", "et_end"])

    before = len(measured)
    measured = measured.drop_duplicates(
        subset=["title_ja", "total_notes", "clear_start_measured"]
    )
    after = len(measured)
    if before != after:
        print(f"  중복 {before - after}곡 제거 → 학습 데이터 {after}곡")

    measured["et_start_ratio"] = measured["et_start"] / measured["total_notes"]
    measured["et_end_ratio"] = measured["et_end"] / measured["total_notes"]

    mean_start = measured.loc[measured["et_start_ratio"] > 0, "et_start_ratio"].mean()
    mean_end = measured.loc[measured["et_end_ratio"] > 0, "et_end_ratio"].mean()

    zero_mask = (measured["et_start_ratio"] == 0) & (measured["et_end_ratio"] == 0)
    n_zero = zero_mask.sum()
    if n_zero > 0:
        print(f"  ET=0/0 곡 {n_zero}개 → 평균 비율로 보정 (시작={mean_start:.3f}, 종료={mean_end:.3f})")
        measured.loc[zero_mask, "et_start_ratio"] = mean_start
        measured.loc[zero_mask, "et_end_ratio"] = mean_end

    X = measured[["total_notes", "et_start_ratio", "et_end_ratio"]]
    y = measured["clear_start_measured"]
    model = LinearRegression().fit(X, y)
    pred = model.predict(X)

    X_simple = measured[["total_notes"]]
    model_simple = LinearRegression().fit(X_simple, y)
    pred_s = model_simple.predict(X_simple)

    return {
        "model": model,
        "model_simple": model_simple,
        "measured": measured,
        "r2": r2_score(y, pred),
        "mae": mean_absolute_error(y, pred),
        "r2_simple": r2_score(y, pred_s),
        "mae_simple": mean_absolute_error(y, pred_s),
        "mean_start": mean_start,
        "mean_end": mean_end,
    }


# ── 전체 예측 ──────────────────────────────────────────────
def predict_all(df: pd.DataFrame, result: dict) -> pd.DataFrame:
    """전체 곡에 대해 예측값 생성."""
    df = df.copy()
    model = result["model"]
    mean_s, mean_e = result["mean_start"], result["mean_end"]

    df["et_start_ratio"] = (df["et_start"] / df["total_notes"]).fillna(0)
    df["et_end_ratio"] = (df["et_end"] / df["total_notes"]).fillna(0)

    zero_mask = (df["et_start_ratio"] == 0) & (df["et_end_ratio"] == 0)
    df.loc[zero_mask, "et_start_ratio"] = mean_s
    df.loc[zero_mask, "et_end_ratio"] = mean_e

    X = df[["total_notes", "et_start_ratio", "et_end_ratio"]]
    df["clear_start_predicted"] = model.predict(X).round().astype(int)
    df["clear_start_measured"] = pd.to_numeric(df["clear_start_measured"], errors="coerce").round().astype("Int64")

    has_et = df["et_start"].notna() & (df["et_start"] > 0)
    df["model_used"] = np.where(has_et, "3변수 OLS", "단순 회귀")
    df["clear_ratio"] = (df["clear_start_predicted"] / df["total_notes"] * 100).round(1)

    return df


# ── 단일 곡 예측 ───────────────────────────────────────────
def predict_one(result: dict, total_notes: int, et_start=None, et_end=None):
    """신곡 1곡 예측."""
    model = result["model"]
    model_simple = result["model_simple"]

    if et_start is not None and et_end is not None and et_start > 0:
        s, e = et_start / total_notes, et_end / total_notes
        X = pd.DataFrame(
            [[total_notes, s, e]],
            columns=["total_notes", "et_start_ratio", "et_end_ratio"],
        )
        pred = model.predict(X)[0]
        used = "3변수 OLS"
    else:
        X = pd.DataFrame([[total_notes]], columns=["total_notes"])
        pred = model_simple.predict(X)[0]
        used = "단순 회귀"

    start_pt = round(pred)
    clear = start_pt + SAFETY_MARGIN
    ratio = start_pt / total_notes * 100
    return start_pt, clear, ratio, used


# ── songs.js 출력 ──────────────────────────────────────────
def export_songs_js(df_pred: pd.DataFrame, result: dict, out_path: str = "songs.js"):
    """
    예측 결과 DataFrame을 songs.js 형식으로 출력.

    출력 파일 구조:
      · SONGS 배열  — voltage_predictor HTML에서 직접 로드
      · MODEL_PARAMS 객체 — 프론트엔드 계산용 모델 파라미터
    """
    model = result["model"]
    model_simple = result["model_simple"]

    # title_ja + total_notes 기준으로 그룹핑 — 합동곡은 units 배열로 수집
    seen: dict = {}
    for _, row in df_pred.iterrows():
        key = (str(row["title_ja"]), int(row["total_notes"]))
        unit = str(row["unit"])
        if key not in seen:
            seen[key] = {"row": row, "units": [unit]}
        else:
            if unit not in seen[key]["units"]:
                seen[key]["units"].append(unit)

    songs = []
    for key, val in seen.items():
        row = val["row"]
        units = val["units"]

        # unitGroup: 첫 번째 유닛 기준 (매핑 없으면 유닛명 그대로)
        primary_unit = units[0]
        unit_group = UNIT_GROUP.get(primary_unit, primary_unit)

        et_s = None if pd.isna(row["et_start"]) else int(row["et_start"])
        et_e = None if pd.isna(row["et_end"]) else int(row["et_end"])
        measured_val = None if pd.isna(row["clear_start_measured"]) else int(row["clear_start_measured"])

        raw_url = row.get("video_url", None)
        if pd.isna(raw_url) or str(raw_url).strip() == "":
            video = None
        else:
            video = str(raw_url).strip()

        duration_str = parse_duration(row.get("duration", None))

        title_ja = key[0]
        title_ja_reading = row.get("title_ja_reading", None)
        title_ko = row.get("title_ko", None)
        title_ko_reading = row.get("title_ko_reading", None)
        title_en = row.get("title_en", None)
        title_en_reading = row.get("title_en_reading", None)

        # video_url_clear: 엑셀 수동 입력값 우선, 없는 예측곡은 duration 절반 지점 자동 생성
        # ※ 미실측 곡의 절반 지점 URL은 메모리(df_pred)에서만 계산되며 엑셀에는 저장되지 않음
        #    → songs.js / CSV에는 반영되지만 엑셀 칼럼은 의도적으로 비워둠
        raw_clear = row.get("video_url_clear", None)
        if pd.notna(raw_clear) and str(raw_clear).strip():
            video_clear = str(raw_clear).strip()
        elif measured_val is None and video:
            video_clear = calc_midpoint_url(video, duration_str)
        else:
            video_clear = None

        predicted_val = int(round(float(row["clear_start_predicted"])))

        song = {
            "type":             str(row["type"]),
            "unit":             " / ".join(units),
            "units":            units,
            "unitGroup":        unit_group,
            "title_ja":         title_ja,
            "title_ja_reading": str(title_ja_reading) if pd.notna(title_ja_reading) else None,
            "title_ko":         str(title_ko) if pd.notna(title_ko) else None,
            "title_ko_reading": str(title_ko_reading) if pd.notna(title_ko_reading) else None,
            "title_en":         str(title_en) if pd.notna(title_en) else None,
            "title_en_reading": str(title_en_reading) if pd.notna(title_en_reading) else None,
            "totalNotes":       key[1],
            "duration":         duration_str,
            "etStart":          et_s,
            "etEnd":            et_e,
            "measured":         measured_val,
            "category":         str(row["category"]),
            "video":            video,
            "videoClear":       video_clear,
            "predicted":        predicted_val,
        }
        if not song.get("title_ko"):
            song["title_ko"] = title_ja
        if not song.get("title_en"):
            song["title_en"] = title_ja
        songs.append(song)

    # 모델 파라미터 (프론트에서 직접 계산할 수 있도록)
    params = {
        "intercept":    model.intercept_,
        "coefs":        list(model.coef_),
        "simpleCoef":   float(model_simple.coef_[0]),
        "simpleInt":    float(model_simple.intercept_),
        "safetyMargin": SAFETY_MARGIN,
        "meanEtStart":  result["mean_start"],
        "meanEtEnd":    result["mean_end"],
        "trainSize":    len(result["measured"]),
        "r2":           result["r2"],
        "mae":          result["mae"],
    }

    songs_json  = json.dumps(songs,  ensure_ascii=False, indent=2)
    params_json = json.dumps(params, ensure_ascii=False, indent=2)

    js_content = f"""\
// songs.js — Voltage Predictor 곡 데이터 & 모델 파라미터
// 총 {len(songs)}곡 | R²={result['r2']:.4f} | MAE={result['mae']:.2f}콤보
// ※ enstars_regression_v3.py 로 자동 생성 — 직접 수정 비권장

const SONGS = {songs_json};

const MODEL_PARAMS = {params_json};
"""

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        f.write(js_content)

    print(f"\n✓ songs.js 저장 완료: {out_path}")
    print(f"  수록곡 {len(songs)}곡 | 실측 {sum(1 for s in songs if s['measured'] is not None)}곡")
    print(f"  모델: R²={result['r2']:.4f}  MAE={result['mae']:.2f}콤보")


# ── 성능 리포트 ────────────────────────────────────────────
def print_report(result: dict):
    r = result
    print(f"\n{'='*55}")
    print(f"  VOLTAGE Easy Clear 예측 모델 v3")
    print(f"{'='*55}")
    print(f"  학습 곡수     : {len(r['measured'])}곡 (중복 제거 후)")
    print(f"  3변수 OLS     : R²={r['r2']:.4f}  MAE={r['mae']:.2f}콤보")
    print(f"  단순 회귀      : R²={r['r2_simple']:.4f}  MAE={r['mae_simple']:.2f}콤보")
    print(f"  SAFETY MARGIN : +{SAFETY_MARGIN}콤보")
    print(f"  평균 ET 비율   : 시작={r['mean_start']:.3f}  종료={r['mean_end']:.3f}")
    print(f"{'='*55}")

    m = r["model"]
    print(f"\n  [3변수 모델 수식]")
    print(f"  시작점 = {m.intercept_:.4f}")
    print(f"         + {m.coef_[0]:.6f} × total_notes")
    print(f"         + ({m.coef_[1]:.4f}) × et_start_ratio")
    print(f"         + {m.coef_[2]:.4f} × et_end_ratio")

    ms = r["model_simple"]
    print(f"\n  [단순 모델 수식]")
    print(f"  시작점 = {ms.coef_[0]:.6f} × total_notes + {ms.intercept_:.4f}")

    measured = r["measured"]
    X = measured[["total_notes", "et_start_ratio", "et_end_ratio"]]
    y = measured["clear_start_measured"].values
    residuals = y - m.predict(X)
    print(f"\n  [잔차 분포]")
    print(f"  평균={residuals.mean():.2f}  표준편차={residuals.std():.2f}")
    print(f"  |오차|≤3 : {(np.abs(residuals) <= 3).sum()}/{len(y)}곡 ({(np.abs(residuals) <= 3).mean()*100:.0f}%)")
    print(f"  |오차|≤5 : {(np.abs(residuals) <= 5).sum()}/{len(y)}곡 ({(np.abs(residuals) <= 5).mean()*100:.0f}%)")


# ── 메인 ───────────────────────────────────────────────────
def main():
    # 인수 파싱
    args = sys.argv[1:]
    do_export = "--export" in args
    args = [a for a in args if a != "--export"]

    out_path = str(_HERE / "songs.js")
    if "--out" in args:
        idx = args.index("--out")
        out_path = args[idx + 1]
        args = [a for a in args if a not in ("--out", out_path)]

    path = args[0] if args else CSV_PATH

    # 로드 & 학습
    print(f"데이터 로드 중: {path}")
    df = load_data(path)
    print(f"  전체 {len(df)}곡 (시트 {df['category'].nunique()}개)")

    print(f"\n  [NaN 확인]")
    print(df[["et_start", "et_end"]].isna().sum())

    result = train_model(df)
    print_report(result)

    df_pred = predict_all(df, result)

    # 미확인 곡 중 video_url_clear가 없으면 duration 절반 지점 URL 자동 생성
    if "video_url_clear" not in df_pred.columns:
        df_pred["video_url_clear"] = None
    mask = (
        df_pred["clear_start_measured"].isna()
        & df_pred["video_url"].notna()
        & (df_pred["video_url"].astype(str).str.strip() != "")
        & (df_pred["video_url_clear"].isna() | (df_pred["video_url_clear"].astype(str).str.strip() == ""))
    )
    df_pred.loc[mask, "video_url_clear"] = df_pred.loc[mask].apply(
        lambda r: calc_midpoint_url(str(r["video_url"]).strip(), parse_duration(r.get("duration"))),
        axis=1,
    )

    # --export 플래그가 있으면 바로 출력하고 종료
    if do_export:
        export_songs_js(df_pred, result, out_path)
        return

    # 전체 예측 결과 표시
    df_show = df_pred.drop_duplicates(subset=["title_ja", "total_notes"])
    cols = [
        "category", "type", "unit", "title_ja", "total_notes", "duration",
        "clear_start_measured", "clear_start_predicted",
        "clear_ratio", "model_used",
    ]
    print(f"\n전체 예측 결과 ({len(df_show)}곡):")
    print(df_show[cols].to_string(index=False))

    # 대화형 메뉴
    print("\n─── 메뉴 ───────────────────────────")
    print("  1) songs.js 저장")
    print("  2) CSV 저장")
    print("  3) 새 곡 예측")
    print("  4) video_url_clear 채우기")
    print("  q) 종료")
    print("────────────────────────────────────")

    while True:
        try:
            cmd = input("선택 (1/2/3/4/q): ").strip().lower()
            if cmd == "q":
                break

            elif cmd == "1":
                export_songs_js(df_pred, result, out_path)

            elif cmd == "2":
                csv_out = path.rsplit(".", 1)[0] + "_result_v3.csv"
                save_cols = [
                    "category", "type", "unit", "title_ja", "title_ja_reading",
                    "title_ko", "title_ko_reading", "title_en", "title_en_reading",
                    "total_notes", "duration", "et_start", "et_end",
                    "clear_start_measured", "clear_start_predicted",
                    "clear_ratio", "model_used", "video_url", "video_url_clear",
                ]
                save_cols = [c for c in save_cols if c in df_show.columns]
                df_show[save_cols].to_csv(csv_out, index=False, encoding="utf-8-sig")
                print(f"✓ CSV 저장: {csv_out}")

            elif cmd == "3":
                print("\n─── 새 곡 예측 (종료: b) ───")
                while True:
                    n_in = input("총 노트수: ").strip()
                    if n_in.lower() == "b":
                        break
                    try:
                        n = int(n_in)
                    except ValueError:
                        print("숫자를 입력해 주세요.")
                        continue
                    s_in = input("앙상블 타임 시작 콤보 (없으면 엔터): ").strip()
                    e_in = input("앙상블 타임 종료 콤보 (없으면 엔터): ").strip()
                    s = int(s_in) if s_in else None
                    e = int(e_in) if e_in else None
                    start_pt, clear, ratio, used = predict_one(result, n, s, e)
                    print(f"  → 시작점: {start_pt}콤보 ({ratio:.1f}%)")
                    print(f"  → 클리어: {clear}콤보  (모델: {used})\n")

            elif cmd == "4":
                api_key = os.getenv("YT_API_KEY", "").strip()
                if not api_key:
                    api_key = input("YouTube API 키 입력: ").strip()
                if not api_key:
                    print("  API 키가 없습니다. .env 파일에 YT_API_KEY를 설정하거나 입력해 주세요.")
                else:
                    filled_map = fill_clear_urls(path, api_key)
                    if filled_map:
                        # df_pred 인메모리 패치
                        mask = df_pred["video_url"].astype(str).str.strip().isin(filled_map)
                        df_pred.loc[mask, "video_url_clear"] = df_pred.loc[mask, "video_url"].map(
                            lambda v: filled_map.get(str(v).strip())
                        )
                        save_yn = input("songs.js / CSV도 저장할까요? (y/n): ").strip().lower()
                        if save_yn == "y":
                            export_songs_js(df_pred, result, out_path)
                            csv_out = path.rsplit(".", 1)[0] + "_result_v3.csv"
                            save_cols = [
                                "category", "type", "unit", "title_ja", "title_ja_reading",
                                "title_ko", "title_ko_reading", "title_en", "title_en_reading",
                                "total_notes", "duration", "et_start", "et_end",
                                "clear_start_measured", "clear_start_predicted",
                                "clear_ratio", "model_used", "video_url", "video_url_clear",
                            ]
                            save_cols = [c for c in save_cols if c in df_pred.columns]
                            df_pred[save_cols].to_csv(csv_out, index=False, encoding="utf-8-sig")
                            print(f"✓ songs.js / CSV 저장 완료")

        except KeyboardInterrupt:
            print("\n종료.")
            break


if __name__ == "__main__":
    main()
